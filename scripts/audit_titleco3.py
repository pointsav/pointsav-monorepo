#!/usr/bin/env python3
"""
audit_titleco3.py — read-only audit of the TitleCo 3 Grande Prairie Excel.

Phase A of the engine-as-canonical migration (per plan dated 2026-06-03):
- Reads all 6 worksheet tabs from the audited Excel
- Cross-references _Proforma / _Report / _CAM tabs per development class
- Emits outputs/audit_titleco3_dev_classes.json as the audit trail

Maps:
  Professional Centres ← Test Site_Proforma + Test Site_Report + Test Site_CAM
  Suburban Office       ← Test Site_Underground_Proforma + Test Site_Underground_Report + Test Site_Underground_CAM

Robust parsing: scans for known labels in each tab and reads the row-wise
values to the right. This avoids hardcoding cell addresses that can shift
between revisions.
"""

from __future__ import annotations
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

ARCHIVE_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ARCHIVE_ROOT / "inputs" / (
    "DUE DILIGENCE_TitleCo 3_2026_01_06_Forecast_Development Proforma_"
    "Test Site_Grande Prairie_FIN.xlsx"
)
OUT_PATH = ARCHIVE_ROOT / "outputs" / "audit_titleco3_dev_classes.json"

CLASS_TABS = {
    "Professional Centres": {
        "proforma": "Test Site_Proforma",
        "report": "Test Site_Report",
        "cam": "Test Site_CAM",
    },
    "Suburban Office": {
        "proforma": "Test Site_Underground_Proforma",
        "report": "Test Site_Underground_Report",
        "cam": "Test Site_Underground_CAM",
    },
}


def row_values(ws, row: int, max_col: int = 16) -> list:
    """Return the values in a row as a list (1-indexed cols 1..max_col)."""
    return [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]


def find_row_by_label(ws, label_substring: str, search_cols=(1, 2, 3, 4), max_row=None) -> int | None:
    """Find the first row where a cell in one of `search_cols` contains `label_substring` (case-insensitive)."""
    if max_row is None:
        max_row = ws.max_row
    needle = label_substring.lower()
    for r in range(1, max_row + 1):
        for c in search_cols:
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and needle in v.lower():
                return r
    return None


def first_numeric(row_vals: list, start_col: int = 1) -> tuple:
    """Find the first numeric value in `row_vals` (0-indexed list) at or after `start_col` (1-indexed)."""
    for i, v in enumerate(row_vals[start_col - 1:], start=start_col):
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return (i, v)
    return (None, None)


def numerics_in_row(row_vals: list, start_col: int = 1) -> list:
    """Return all (col_1idx, value) pairs where the cell is numeric and > 0."""
    out = []
    for i, v in enumerate(row_vals[start_col - 1:], start=start_col):
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append((i, v))
    return out


def audit_proforma(ws) -> dict:
    """Pull rental area summary, totals, and performance measures."""
    # Rental area summary: find "Rental area summary" header row, then read rows below for components
    header_row = find_row_by_label(ws, "Rental area summary")
    rental_area = []
    total_nla = None
    total_rent = None
    if header_row:
        # Component rows typically follow immediately (Underground, Retail, Office Floor 1-6)
        # Stop at the "Totals" row
        for r in range(header_row + 1, header_row + 12):
            row_vals = row_values(ws, r, 16)
            label = None
            for c_idx in (3, 4):  # label may be in col C or D
                v = ws.cell(row=r, column=c_idx).value
                if v and isinstance(v, str) and v.strip():
                    label = v.strip()
                    break
            if not label:
                continue
            # Look for "Totals" row
            if "total" in label.lower():
                # Totals row — read directly from stable column positions:
                #   col F (6) = unit count; col G (7) = NLA sqft; col J (10) = rent at lease start
                total_nla = ws.cell(row=r, column=7).value
                total_rent = ws.cell(row=r, column=10).value
                # Fallback if col 10 missing — find largest numeric ignoring col B (line number)
                if total_rent is None:
                    nums = numerics_in_row(row_vals, start_col=5)  # skip line-number col B
                    if nums:
                        total_rent = max(nums, key=lambda x: x[1])[1]
                break
            # Component row — read from stable column positions (col F-K = 6-11)
            sqft = ws.cell(row=r, column=7).value  # col G
            rate = ws.cell(row=r, column=8).value  # col H
            annual = ws.cell(row=r, column=10).value  # col J (rent at lease start)
            if sqft and isinstance(sqft, (int, float)) and sqft > 0:
                rental_area.append({
                    "component": label,
                    "nla_sqft": sqft,
                    "rate_per_sqft": rate,
                    "annual_rent": annual,
                })

    # Total project cost — find label
    tpc_row = find_row_by_label(ws, "TOTAL COSTS ($)", search_cols=(1, 2, 3, 4))
    total_project_cost = None
    cost_per_sqft = None
    if tpc_row:
        _, total_project_cost = first_numeric(row_values(ws, tpc_row, 16), start_col=3)
        # Cost per sf is the next row
        _, cost_per_sqft = first_numeric(row_values(ws, tpc_row + 1, 16), start_col=3)

    # Development yield + Net initial yield
    dy_row = find_row_by_label(ws, "Development yield", search_cols=(1, 2, 3, 4))
    dev_yield = None
    if dy_row:
        _, dev_yield = first_numeric(row_values(ws, dy_row, 16), start_col=3)
    niy_row = find_row_by_label(ws, "Net initial yield", search_cols=(1, 2, 3, 4))
    net_initial_yield = None
    if niy_row:
        _, net_initial_yield = first_numeric(row_values(ws, niy_row, 16), start_col=3)

    # GDV (Investment valuations totals): find "Investment valuations", then the "Totals" row + any "Capitalized rent" rows.
    iv_row = find_row_by_label(ws, "Investment valuations", search_cols=(1, 2, 3, 4))
    gdv = None
    capitalization_rate_from_valuation = None
    capitalization_components = []
    if iv_row:
        for r in range(iv_row + 1, iv_row + 15):
            # Labels in this section live in cols C, D, OR E (sub-indented)
            label_cell = None
            for cc in (3, 4, 5):
                v = ws.cell(row=r, column=cc).value
                if v and isinstance(v, str) and v.strip():
                    label_cell = v.strip()
                    break
            if not label_cell:
                continue
            ll = label_cell.lower()
            if "total" in ll and gdv is None:
                # Totals row — GDV is the first numeric (skip line-number col B)
                nums = numerics_in_row(row_values(ws, r, 16), start_col=4)
                if nums:
                    gdv = nums[0][1]
                continue
            if "capitalized rent" in ll:
                # cols: col F (6) sometimes carries net rent; col H (8) carries cap rate; col I (9) carries this component's GDV
                rate_cell = ws.cell(row=r, column=8).value
                gdv_cell = ws.cell(row=r, column=9).value
                # net rent might be on this row col F/G, or on the previous row (5.500% non-recovery cost) col G (7)
                net_rent_cell = ws.cell(row=r, column=7).value
                if net_rent_cell is None:
                    # Look at the row above
                    prev = ws.cell(row=r - 1, column=7).value
                    if isinstance(prev, (int, float)):
                        net_rent_cell = prev
                if isinstance(gdv_cell, (int, float)) and isinstance(net_rent_cell, (int, float)) and gdv_cell > 0 and net_rent_cell > 0:
                    rate_back = net_rent_cell / gdv_cell
                    capitalization_components.append({
                        "label_row_above": label_cell,
                        "net_rent": net_rent_cell,
                        "gdv_component": gdv_cell,
                        "rate_displayed": rate_cell,
                        "rate_back_computed": rate_back,
                    })
                    # Trust back-computed (Excel rounds display)
                    capitalization_rate_from_valuation = rate_back

    return {
        "rental_area_components": rental_area,
        "total_nla_sqft": total_nla,
        "total_rent_at_lease_start": total_rent,
        "total_project_cost": total_project_cost,
        "cost_per_sqft_gross": cost_per_sqft,
        "development_yield_on_rents": dev_yield,
        "net_initial_yield_excel_display": net_initial_yield,
        "capitalization_rate_from_valuation": capitalization_rate_from_valuation,
        "capitalization_components": capitalization_components,
        "gross_development_value": gdv,
    }


def audit_report(ws) -> dict:
    """Pull cost stack from Report tab. Categories are in col B/C with rate in col D and pct in col E."""
    # Find "Total project costs" row as the end-marker; iterate rows above it
    end_row = find_row_by_label(ws, "Total project costs", search_cols=(1, 2, 3, 4))
    categories = []
    if end_row:
        for r in range(9, end_row + 1):  # rows 9 to end_row inclusive
            label = ws.cell(row=r, column=3).value
            if not label or not isinstance(label, str) or not label.strip():
                continue
            cost_per_sf = ws.cell(row=r, column=4).value
            pct_of_total = ws.cell(row=r, column=5).value
            if cost_per_sf is None:
                continue
            categories.append({
                "category": label.strip(),
                "cost_per_sqft_gross": cost_per_sf,
                "pct_of_total": pct_of_total,
            })
    return {"cost_stack": categories}


def audit_cam(ws) -> dict:
    """Pull operating expenses from CAM tab. Label col C, NLA ref col G, rate col H, cost col I."""
    # Find "EXPENSES" header
    exp_header = find_row_by_label(ws, "EXPENSES", search_cols=(1, 2, 3, 4))
    end_row = find_row_by_label(ws, "Total expenses", search_cols=(1, 2, 3, 4))
    expenses = []
    if exp_header and end_row:
        for r in range(exp_header + 1, end_row + 1):
            label = ws.cell(row=r, column=3).value
            if not label or not isinstance(label, str) or not label.strip():
                continue
            nla_ref = ws.cell(row=r, column=7).value
            rate = ws.cell(row=r, column=8).value
            cost = ws.cell(row=r, column=9).value
            if cost is None:
                continue
            if "total" in label.lower():
                continue  # captured separately as total_annual_expenses
            expenses.append({
                "category": label.strip(),
                "nla_reference_sqft": nla_ref,
                "rate": rate,
                "annual_cost": cost,
            })

    total_expenses = None
    if end_row:
        total_expenses = ws.cell(row=end_row, column=9).value

    # NOI line — "Net operating income"
    noi_row = find_row_by_label(ws, "Net operating income", search_cols=(1, 2, 3, 4))
    noi = None
    if noi_row:
        _, noi = first_numeric(row_values(ws, noi_row, 16), start_col=4)

    # Gross rental income
    gri_row = find_row_by_label(ws, "Gross rental income", search_cols=(1, 2, 3, 4))
    gri = None
    if gri_row:
        _, gri = first_numeric(row_values(ws, gri_row, 16), start_col=4)

    return {
        "operating_expenses": expenses,
        "total_annual_expenses": total_expenses,
        "gross_rental_income": gri,
        "net_operating_income": noi,
    }


def calibrate(audit: dict, target_dev_yield: float = 0.105, target_cap_rate: float = 0.0625) -> dict:
    """Compute required rent at target dev yield + GDV at target cap rate."""
    total_cost = audit["proforma"]["total_project_cost"]
    current_rent = audit["proforma"]["total_rent_at_lease_start"]
    current_nla = audit["proforma"]["total_nla_sqft"]
    if total_cost is None or current_rent is None or current_nla in (None, 0):
        return {"calibration_error": "missing input values",
                "total_cost": total_cost, "current_rent": current_rent, "current_nla": current_nla}

    required_rent = target_dev_yield * total_cost
    rent_uplift = required_rent - current_rent
    uplift_pct = (rent_uplift / current_rent) * 100.0 if current_rent else None
    rent_uplift_per_sf = rent_uplift / current_nla

    net_rent = required_rent * (1.0 - 0.055)
    gdv_at_target = net_rent / target_cap_rate

    return {
        "target_dev_yield": target_dev_yield,
        "target_cap_rate": target_cap_rate,
        "current_dev_yield": audit["proforma"]["development_yield_on_rents"],
        "current_cap_rate": audit["proforma"]["capitalization_rate_from_valuation"],
        "current_rent": current_rent,
        "required_rent_at_target_yield": required_rent,
        "rent_uplift_absolute": rent_uplift,
        "rent_uplift_percent": uplift_pct,
        "rent_uplift_per_sqft_nla": rent_uplift_per_sf,
        "current_rent_per_sqft_nla": current_rent / current_nla,
        "required_rent_per_sqft_nla": required_rent / current_nla,
        "net_rent_after_5_5_pct_non_recovery": net_rent,
        "gdv_at_target_cap_rate": gdv_at_target,
    }


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: Excel input not found: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading: {XLSX_PATH.name}")
    wb = load_workbook(XLSX_PATH, data_only=True)
    print(f"  Worksheets: {wb.sheetnames}\n")

    audit_output = {
        "schema_version": "audit-titleco3-v1",
        "source_xlsx": XLSX_PATH.name,
        "audit_date": "2026-06-03",
        "operator_brief_reference": "briefs/BRIEF-tool-proforma-leapfrog-2030.md v0.15.9 §5h",
        "target_calibration": {
            "development_yield": 0.105,
            "cap_rate": 0.0625,
        },
        "classes": {},
    }

    for class_name, tab_map in CLASS_TABS.items():
        print(f"--- {class_name} ---")
        cls = {"tabs": tab_map}
        proforma_ws = wb[tab_map["proforma"]]
        report_ws = wb[tab_map["report"]]
        cam_ws = wb[tab_map["cam"]]

        cls["proforma"] = audit_proforma(proforma_ws)
        cls["report"] = audit_report(report_ws)
        cls["cam"] = audit_cam(cam_ws)

        proforma_nla = cls["proforma"]["total_nla_sqft"]
        proforma_tpc = cls["proforma"]["total_project_cost"]
        proforma_rent = cls["proforma"]["total_rent_at_lease_start"]
        proforma_dy = cls["proforma"]["development_yield_on_rents"]
        proforma_cr = cls["proforma"]["capitalization_rate_from_valuation"]

        print(f"  Total NLA: {proforma_nla:,.0f} sf" if proforma_nla else "  Total NLA: NOT FOUND")
        print(f"  Total project cost: ${proforma_tpc:,.2f}" if proforma_tpc else "  Total project cost: NOT FOUND")
        print(f"  Cost per sf gross: ${cls['proforma']['cost_per_sqft_gross']:,.2f}" if cls['proforma']['cost_per_sqft_gross'] else "  Cost per sf: NOT FOUND")
        print(f"  Current rent: ${proforma_rent:,.2f}" if proforma_rent else "  Current rent: NOT FOUND")
        print(f"  Dev yield (Excel): {proforma_dy:.4f}" if proforma_dy else "  Dev yield: NOT FOUND")
        print(f"  Cap rate (back-computed from valuation): {proforma_cr:.4f}" if proforma_cr else "  Cap rate: NOT FOUND")
        print(f"  NIY (Excel display row 94): {cls['proforma']['net_initial_yield_excel_display']:.4f}" if cls['proforma']['net_initial_yield_excel_display'] else "")
        print(f"  GDV (Excel): ${cls['proforma']['gross_development_value']:,.2f}" if cls['proforma']['gross_development_value'] else "")
        print(f"  Cost-stack categories found: {len(cls['report']['cost_stack'])}")
        print(f"  CAM expense lines found: {len(cls['cam']['operating_expenses'])}")
        print(f"  CAM total expenses: ${cls['cam']['total_annual_expenses']:,.2f}" if cls['cam']['total_annual_expenses'] else "")
        print(f"  CAM NOI: ${cls['cam']['net_operating_income']:,.2f}" if cls['cam']['net_operating_income'] else "")

        cls["calibration"] = calibrate(cls)
        cal = cls["calibration"]
        if "required_rent_at_target_yield" in cal:
            print(f"  → Required rent at 10.5%: ${cal['required_rent_at_target_yield']:,.2f}")
            print(f"  → Rent uplift: ${cal['rent_uplift_absolute']:,.2f} ({cal['rent_uplift_percent']:.2f}%)")
            print(f"  → GDV at 6.25% cap: ${cal['gdv_at_target_cap_rate']:,.2f}")

        audit_output["classes"][class_name] = cls
        print()

    # Tech Industrial + Retail Select — no Excel; flagged
    for class_name in ("Tech Industrial", "Retail Select"):
        print(f"--- {class_name} ---")
        print(f"  NO Excel proforma exists in this workbook.")
        print(f"  Engine extrapolation per plan (Phase B); see brief §5h class composition.")
        audit_output["classes"][class_name] = {
            "no_excel_source": True,
            "note": "No per-class Excel proforma in this workbook. Engine extrapolation per plan Phase B.",
            "brief_class_total_nla": (459_630 if class_name == "Tech Industrial" else 229_815),
            "brief_building_count_or_pairs": (
                "30 pairs (60 buildings) — each pair = 7,200 + 8,400 sf"
                if class_name == "Tech Industrial"
                else "18 pair-structures (36 buildings) — variant mix of 4,500/6,700/7,700 sf paired up"
            ),
        }
        print()

    audit_output["portfolio_rollup_check"] = {
        "source": "DUE DILIGENCE_MCorp_Tear Sheet_Alternative Real Estate_FIN.xlsx",
        "total_portfolio_nla": 2_298_150,
        "total_project_cost": 750_000_000,
        "blended_cost_per_sqft_nla": 326.35,
        "implied_portfolio_noi_at_10_5_pct_yield": 78_750_000,
        "implied_portfolio_gdv_at_6_25_pct_cap": 1_260_000_000,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w") as f:
        json.dump(audit_output, f, indent=2, default=str)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
