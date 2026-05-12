"""
Fill INVOICE YEAR (col C) and INVOICE QTR (col D) for all data rows
that have an INVOICE DATE but no YEAR/QTR value.

Header row: 10, Sheet: Sheet1.
Usage:  python3 fill_year_qtr.py <input.xlsx> <year> <qtr>
Example: python3 fill_year_qtr.py "April Transactions_WEEKENDS_FILLED.xlsx" 2026 Q2
Output: <stem>_YEARQTR.xlsx
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

SHEET       = "Sheet1"
HEADER_ROW  = 10
DATE_HEADER = "INVOICE DATE"
YEAR_HEADER = "INVOICE YEAR"
QTR_HEADER  = "INVOICE QTR"
OUTPUT_SUFFIX = "_YEARQTR"


def find_column(ws, header_row: int, header_text: str) -> int:
    target = header_text.strip().casefold()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if isinstance(v, str) and v.strip().casefold() == target:
            return col
    raise ValueError(f"Header {header_text!r} not found on row {header_row}")


def main():
    if len(sys.argv) != 4:
        print("Usage: python3 fill_year_qtr.py <input.xlsx> <year> <qtr>")
        print("  e.g. python3 fill_year_qtr.py 'April Transactions_WEEKENDS_FILLED.xlsx' 2026 Q2")
        sys.exit(1)

    in_path   = Path(sys.argv[1])
    year_val  = int(sys.argv[2])
    qtr_val   = sys.argv[3]
    out_path  = in_path.with_name(in_path.stem + OUTPUT_SUFFIX + in_path.suffix)

    print(f"Reading:  {in_path}")
    print(f"Writing:  {out_path}")
    print(f"Values:   INVOICE YEAR={year_val}  INVOICE QTR={qtr_val}")

    wb = load_workbook(in_path)
    ws = wb[SHEET]

    date_col = find_column(ws, HEADER_ROW, DATE_HEADER)
    year_col = find_column(ws, HEADER_ROW, YEAR_HEADER)
    qtr_col  = find_column(ws, HEADER_ROW, QTR_HEADER)
    print(f"Columns:  DATE={date_col}  YEAR={year_col}  QTR={qtr_col}")

    filled = skipped = already_set = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        date_val = ws.cell(row=row, column=date_col).value
        if date_val is None:
            skipped += 1
            continue
        existing_year = ws.cell(row=row, column=year_col).value
        existing_qtr  = ws.cell(row=row, column=qtr_col).value
        if existing_year is not None or existing_qtr is not None:
            already_set += 1
            continue
        ws.cell(row=row, column=year_col).value = year_val
        ws.cell(row=row, column=qtr_col).value  = qtr_val
        filled += 1

    wb.save(out_path)
    print(f"\nRows filled:      {filled}")
    print(f"Rows already set: {already_set}")
    print(f"Rows skipped (no date): {skipped}")
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
