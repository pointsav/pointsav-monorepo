"""
Woodfine Management Corp. — Monthly Bookkeeping Pipeline
=========================================================
Processes a new monthly transactions file and appends it to the master ledger.

Steps (in order):
  1  Fix text-formatted dates (MM/DD/YYYY → Excel datetime)
  2  Fill WEEKENDS column  (BC + federal holiday calendar)
  3  Fill INVOICE YEAR and INVOICE QTR  (auto-detected from the data)
  4  Extract foreign currency from PAYEE column
  5  Fill COMPANY, TYPE, CATEGORY from Data_2026 payee lookup
  6  Standardise BANK names  (TD → TD - JMW, TD Visa → TD - Visa)
  7  Detect and remove duplicate rows  (matched by INVOICE DATE + CDN amount)
  8  Write monthly summary CSV
  9  Append clean rows to the master ledger (safe copy)

Usage:
    python3 run_monthly_pipeline.py <new_transactions.xlsx> <master_ledger.xlsx>

Example:
    python3 run_monthly_pipeline.py "May Transactions.xlsx" "Data_2026.xlsx"

Output files (written alongside the input file):
    <stem>_PROCESSED.xlsx          — fully processed transactions (all steps applied)
    <stem>_SUMMARY.csv             — monthly summary report
    <master>_updated.xlsx          — master ledger safe copy with new rows appended

The original input file and master ledger are never modified.
"""

import re
import sys
import csv
import shutil
from datetime import date, timedelta, datetime
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════ #
#  Schema constants — update these if the ledger layout ever changes
# ═══════════════════════════════════════════════════════════════════════════ #
SHEET       = "Sheet1"
HEADER_ROW  = 10

C_REF       =  2   # REF NUMBER
C_YEAR      =  3   # INVOICE YEAR
C_QTR       =  4   # INVOICE QTR
C_DATE      =  5   # INVOICE DATE
C_TRAVEL    =  6   # TRAVEL DAYS
C_WEEKENDS  =  7   # WEEKENDS
C_COMPANY   =  8   # COMPANY
C_PAYEE     =  9   # PAYEE
C_TYPE      = 10   # TYPE
C_CATEGORY  = 11   # CATEGORY
C_ATTENDEE  = 12   # ATTENDEE
C_DESC      = 13   # DESCRIPTION
C_FX_AMT   = 18   # FORIGN CURRENCY AMOUNT
C_FX_CUR   = 19   # Curency
C_FX_RATE  = 20   # Exchange rate
C_CDN      = 21   # CDN
C_BANK     = 22   # BANK
C_ACCOUNT  = 23   # ACCOUNT
C_RECEIPT  = 24   # RECEIPT

BANK_RENAMES = {
    "TD":      "TD - JMW",
    "TD Visa": "TD - Visa",
}

FX_PATTERN = re.compile(
    r"(?P<amt>[-+]?\d*\.?\d+)\s*(?P<cur>[A-Z]{3})\s*@\s*(?P<rate>[-+]?\d*\.?\d+)"
)

# ═══════════════════════════════════════════════════════════════════════════ #
#  Holiday calendar  (BC + federal)
# ═══════════════════════════════════════════════════════════════════════════ #
def _easter(year):
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)

def _nth_weekday(year, month, weekday, n):
    first  = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (n - 1))

def _holidays(year):
    easter  = _easter(year)
    may24   = date(year, 5, 24)
    vic_day = may24 - timedelta(days=may24.weekday() % 7)
    return {
        date(year, 1,  1):              "New Year's Day",
        _nth_weekday(year, 2, 0, 3):    "Family Day",
        easter - timedelta(days=2):     "Good Friday",
        vic_day:                        "Victoria Day",
        date(year, 7,  1):              "Canada Day",
        _nth_weekday(year, 8, 0, 1):    "BC Day",
        _nth_weekday(year, 9, 0, 1):    "Labour Day",
        date(year, 9, 30):              "Truth and Reconciliation Day",
        _nth_weekday(year, 10, 0, 2):   "Thanksgiving",
        date(year, 11, 11):             "Remembrance Day",
        date(year, 12, 25):             "Christmas Day",
        date(year, 12, 26):             "Boxing Day",
    }

def _classify(d, holidays):
    if d in holidays:
        return "Holiday"
    if d.weekday() >= 5:
        return "Weekend"
    return "Business Day"

def _month_to_qtr(month):
    return f"Q{(month - 1) // 3 + 1}"

def _parse_text_date(raw):
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            pass
    return None

def _to_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, date):
        return v
    return None

# ═══════════════════════════════════════════════════════════════════════════ #
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════ #
def _banner(step, title):
    print(f"\n{'─'*60}")
    print(f"  Step {step}: {title}")
    print(f"{'─'*60}")

def _data_rows(ws):
    """Yield (row_number, row_tuple) for every non-blank row after the header."""
    for i, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1
    ):
        if any(v is not None for v in row):
            yield i, row

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 1 — Fix text-formatted dates
# ═══════════════════════════════════════════════════════════════════════════ #
def step1_fix_text_dates(ws, all_holidays):
    fixed = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=C_DATE)
        val  = cell.value
        if val is None or hasattr(val, "year"):
            continue
        d = _parse_text_date(str(val))
        if d is None:
            print(f"    WARNING row {row}: cannot parse date {val!r} — left as-is")
            continue
        cell.value = datetime(d.year, d.month, d.day)
        fixed += 1
    print(f"  Text dates converted: {fixed}")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 2 — Fill WEEKENDS
# ═══════════════════════════════════════════════════════════════════════════ #
def step2_fill_weekends(ws, all_holidays):
    filled = skipped = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        date_val = ws.cell(row=row, column=C_DATE).value
        if not hasattr(date_val, "year"):
            skipped += 1
            continue
        d = date_val.date() if hasattr(date_val, "date") else date_val
        ws.cell(row=row, column=C_WEEKENDS).value = _classify(d, all_holidays)
        filled += 1
    dist = Counter(
        ws.cell(row=r, column=C_WEEKENDS).value
        for r in range(HEADER_ROW + 1, ws.max_row + 1)
        if ws.cell(row=r, column=C_WEEKENDS).value
    )
    print(f"  WEEKENDS filled: {filled}  skipped (no date): {skipped}")
    print(f"  Distribution: {dict(dist)}")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 3 — Fill INVOICE YEAR and QTR  (auto-detect)
# ═══════════════════════════════════════════════════════════════════════════ #
def step3_fill_year_qtr(ws):
    months = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=C_DATE).value
        d = _to_date(v) if hasattr(v, "year") else None
        if d:
            months.append((d.year, d.month))

    if not months:
        print("  WARNING: no valid dates found — YEAR/QTR not filled")
        return

    # Primary year = most common year; primary month = most common month in that year
    primary_year = Counter(y for y, _ in months).most_common(1)[0][0]
    primary_month = Counter(m for y, m in months if y == primary_year).most_common(1)[0][0]
    primary_qtr = _month_to_qtr(primary_month)
    print(f"  Auto-detected primary period: {primary_year} {primary_qtr}")

    filled = corrected = skipped = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=C_DATE).value
        d = _to_date(v) if hasattr(v, "year") else None
        if d is None:
            skipped += 1
            continue
        yr_cell  = ws.cell(row=row, column=C_YEAR)
        qtr_cell = ws.cell(row=row, column=C_QTR)
        new_qtr  = _month_to_qtr(d.month)
        if yr_cell.value is None and qtr_cell.value is None:
            yr_cell.value  = d.year
            qtr_cell.value = new_qtr
            if new_qtr != primary_qtr:
                corrected += 1
                print(f"    Row {row}: {d} → {d.year} {new_qtr}  (different quarter — spill-over row)")
            else:
                filled += 1
    print(f"  YEAR/QTR filled: {filled + corrected}  "
          f"(primary {primary_qtr}: {filled}, other quarter: {corrected}, "
          f"skipped no-date: {skipped})")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 4 — Extract foreign currency from PAYEE
# ═══════════════════════════════════════════════════════════════════════════ #
def step4_extract_fx(ws):
    mutations = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        src_cell = ws.cell(row=row, column=C_PAYEE)
        val = src_cell.value
        if not isinstance(val, str):
            continue
        m = FX_PATTERN.search(val)
        if not m:
            continue
        rate_raw = m.group("rate")
        if rate_raw.startswith((".", "-.", "+.")):
            rate_raw = rate_raw.replace(".", "0.", 1)
        ws.cell(row=row, column=C_FX_AMT).value  = float(m.group("amt"))
        ws.cell(row=row, column=C_FX_CUR).value  = m.group("cur")
        ws.cell(row=row, column=C_FX_RATE).value = float(rate_raw)
        clean = (val[:m.start()] + val[m.end():]).strip()
        src_cell.value = re.sub(r"\s{2,}", " ", clean)
        mutations += 1
    print(f"  Payee rows with FX extracted: {mutations}")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 5 + 6 — Fill lookup fields and rename banks
# ═══════════════════════════════════════════════════════════════════════════ #
def _build_payee_lookup(master_path):
    wb = load_workbook(master_path, read_only=True, data_only=True)
    ws = wb.active
    raw = defaultdict(lambda: {"COMPANY": set(), "TYPE": set(), "CATEGORY": set()})
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        payee = row[C_PAYEE - 1]
        if payee is None:
            continue
        key = str(payee).strip().upper()
        for field, col in [("COMPANY", C_COMPANY), ("TYPE", C_TYPE), ("CATEGORY", C_CATEGORY)]:
            val = row[col - 1]
            if val is not None:
                raw[key][field].add(str(val).strip())
    wb.close()
    return {
        key: {f: list(v)[0] for f, v in fields.items() if len(v) == 1}
        for key, fields in raw.items()
        if any(len(v) == 1 for v in fields.values())
    }

def step5_fill_lookup(ws, lookup):
    filled = defaultdict(int)
    no_match = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        payee = ws.cell(row=row, column=C_PAYEE).value
        if payee is None:
            continue
        entry = lookup.get(str(payee).strip().upper(), {})
        if not entry:
            no_match += 1
            continue
        for field, col in [("COMPANY", C_COMPANY), ("TYPE", C_TYPE), ("CATEGORY", C_CATEGORY)]:
            cell = ws.cell(row=row, column=col)
            if cell.value is None and field in entry:
                cell.value = entry[field]
                filled[field] += 1
    for field in ["COMPANY", "TYPE", "CATEGORY"]:
        print(f"  {field:<12} filled in {filled[field]} rows")
    print(f"  No match in master ledger: {no_match} rows")

def step6_rename_banks(ws):
    renamed = defaultdict(int)
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=C_BANK)
        if cell.value is not None:
            old = str(cell.value).strip()
            new = BANK_RENAMES.get(old)
            if new:
                cell.value = new
                renamed[old] += 1
    for old, count in renamed.items():
        print(f"  {old!r} → {BANK_RENAMES[old]!r}  ({count} rows)")
    if not renamed:
        print("  No bank names needed renaming")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 7 — Remove duplicates vs master ledger
# ═══════════════════════════════════════════════════════════════════════════ #
def _master_keys(master_path):
    wb = load_workbook(master_path, read_only=True, data_only=True)
    ws = wb.active
    keys = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        d = _to_date(row[C_DATE - 1]) if hasattr(row[C_DATE - 1], "year") else None
        c = row[C_CDN - 1]
        if d is not None and c is not None:
            keys.add((d, c))
    wb.close()
    return keys

def step7_remove_duplicates(ws, master_path):
    existing = _master_keys(master_path)
    rows_to_delete = []
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=C_DATE).value
        d = _to_date(v) if hasattr(v, "year") else None
        c = ws.cell(row=row, column=C_CDN).value
        if d is None and c is None:
            continue
        if (d, c) in existing:
            payee = ws.cell(row=row, column=C_PAYEE).value
            rows_to_delete.append((row, d, c, payee))

    if rows_to_delete:
        print(f"  Removing {len(rows_to_delete)} rows already in master ledger:")
        for r, d, c, payee in rows_to_delete:
            print(f"    row {r}: date={d}  CDN={c}  payee={payee}")
        for r, _, _, _ in sorted(rows_to_delete, key=lambda x: x[0], reverse=True):
            ws.delete_rows(r)
    else:
        print("  No duplicates found")

    return len(rows_to_delete)

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 8 — Monthly summary CSV
# ═══════════════════════════════════════════════════════════════════════════ #
def step8_summary(ws, out_path):
    total_rows = total_cdn = total_debit = total_credit = 0
    bank_counts = defaultdict(int)
    bank_cdn    = defaultdict(float)
    weekends    = defaultdict(int)
    fx_count    = 0
    currency_counts = defaultdict(int)

    for _, row in _data_rows(ws):
        total_rows += 1
        cdn = row[C_CDN - 1]
        bank = str(row[C_BANK - 1]).strip() if row[C_BANK - 1] else "(none)"
        bank_counts[bank] += 1
        if cdn is not None:
            try:
                amt = float(cdn)
                total_cdn   += amt
                total_debit  += amt if amt >= 0 else 0
                total_credit += amt if amt < 0 else 0
                bank_cdn[bank] += amt
            except (TypeError, ValueError):
                pass
        wk = row[C_WEEKENDS - 1]
        weekends[str(wk) if wk else "(not set)"] += 1
        if row[C_FX_AMT - 1] is not None:
            fx_count += 1
            cur = str(row[C_FX_CUR - 1]).strip() if row[C_FX_CUR - 1] else "(unknown)"
            currency_counts[cur] += 1

    rows = [
        ["SECTION", "LABEL", "VALUE"], ["", "", ""],
        ["TOTALS", "Transaction rows",   total_rows],
        ["TOTALS", "Net CDN amount",     f"{total_cdn:,.2f}"],
        ["TOTALS", "Total debits (>=0)", f"{total_debit:,.2f}"],
        ["TOTALS", "Total credits (<0)", f"{total_credit:,.2f}"],
        ["", "", ""],
        ["BY BANK", "Bank", "Rows | CDN net"],
    ]
    for bank in sorted(bank_counts):
        rows.append(["BY BANK", bank, f"{bank_counts[bank]} rows | {bank_cdn[bank]:,.2f} CDN"])
    rows += [["", "", ""], ["WEEKENDS", "Classification", "Count"]]
    for label in sorted(weekends):
        rows.append(["WEEKENDS", label, weekends[label]])
    rows += [["", "", ""], ["FX", "Foreign currency rows", fx_count]]
    for cur in sorted(currency_counts):
        rows.append(["FX", f"  {cur}", currency_counts[cur]])

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)

    print(f"  Rows: {total_rows}  |  Net CDN: {total_cdn:,.2f}"
          f"  |  Debits: {total_debit:,.2f}  Credits: {total_credit:,.2f}")
    print(f"  FX rows: {fx_count}  |  Summary written to: {out_path.name}")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Step 9 — Append to master ledger (safe copy)
# ═══════════════════════════════════════════════════════════════════════════ #
def step9_append(processed_path, master_path, out_path):
    shutil.copy2(master_path, out_path)

    wb_src = load_workbook(processed_path, read_only=True, data_only=True)
    ws_src = wb_src.active
    new_rows = [
        row for row in ws_src.iter_rows(min_row=HEADER_ROW + 1, values_only=True)
        if any(v is not None for v in row)
    ]
    wb_src.close()

    wb_dst = load_workbook(out_path)
    ws_dst = wb_dst.active

    last = HEADER_ROW
    for row in ws_dst.iter_rows(min_row=HEADER_ROW + 1):
        for cell in row:
            if cell.value is not None:
                last = cell.row
                break

    for i, row_vals in enumerate(new_rows):
        dest = last + 1 + i
        for col_idx, val in enumerate(row_vals, 1):
            if val is not None:
                ws_dst.cell(row=dest, column=col_idx, value=val)

    wb_dst.save(out_path)
    before = last - HEADER_ROW
    print(f"  Master ledger rows before: {before}")
    print(f"  Rows appended:             {len(new_rows)}")
    print(f"  Total rows after:          {before + len(new_rows)}")
    print(f"  Safe copy written to:      {out_path.name}")

# ═══════════════════════════════════════════════════════════════════════════ #
#  Main
# ═══════════════════════════════════════════════════════════════════════════ #
def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    input_path  = Path(sys.argv[1])
    master_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}")
        sys.exit(1)
    if not master_path.exists():
        print(f"ERROR: master ledger not found: {master_path}")
        sys.exit(1)

    processed_path = input_path.with_name(input_path.stem + "_PROCESSED.xlsx")
    summary_path   = input_path.with_name(input_path.stem + "_SUMMARY.csv")
    updated_path   = master_path.with_name(master_path.stem + "_updated.xlsx")

    print(f"\nWOODFINE MONTHLY BOOKKEEPING PIPELINE")
    print(f"Input:   {input_path.name}")
    print(f"Master:  {master_path.name}")
    print(f"Output:  {processed_path.name}")

    # Copy input → processed (we work on the copy; original stays clean)
    shutil.copy2(input_path, processed_path)
    wb = load_workbook(processed_path)
    ws = wb[SHEET]

    # Build holiday calendar for a span of years
    all_holidays = {}
    for yr in range(2023, 2030):
        all_holidays.update(_holidays(yr))

    _banner(1, "Fix text-formatted dates")
    step1_fix_text_dates(ws, all_holidays)

    _banner(2, "Fill WEEKENDS")
    step2_fill_weekends(ws, all_holidays)

    _banner(3, "Fill INVOICE YEAR and QTR  (auto-detected)")
    step3_fill_year_qtr(ws)

    _banner(4, "Extract foreign currency from PAYEE")
    step4_extract_fx(ws)

    _banner(5, "Fill COMPANY / TYPE / CATEGORY from master ledger")
    lookup = _build_payee_lookup(master_path)
    step5_fill_lookup(ws, lookup)

    _banner(6, "Standardise BANK names")
    step6_rename_banks(ws)

    _banner(7, "Remove duplicates vs master ledger")
    removed = step7_remove_duplicates(ws, master_path)

    # Save processed file before summary (summary reads from it)
    wb.save(processed_path)

    _banner(8, "Monthly summary")
    # Reload to get accurate counts after row deletions
    wb2 = load_workbook(processed_path, read_only=True, data_only=True)
    ws2 = wb2[SHEET]
    step8_summary(ws2, summary_path)
    wb2.close()

    _banner(9, "Append to master ledger")
    step9_append(processed_path, master_path, updated_path)

    print(f"\n{'═'*60}")
    print(f"  PIPELINE COMPLETE")
    print(f"{'═'*60}")
    print(f"  Processed file:  {processed_path.name}")
    print(f"  Summary CSV:     {summary_path.name}")
    print(f"  Updated ledger:  {updated_path.name}  ← review then rename to {master_path.name}")
    print(f"\n  Rows removed as duplicates: {removed}")
    print(f"  Original input file unchanged: {input_path.name}")
    print(f"  Original master unchanged:     {master_path.name}")


if __name__ == "__main__":
    main()
