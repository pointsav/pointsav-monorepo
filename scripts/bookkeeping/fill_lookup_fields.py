"""
Two operations on the April FINAL file:

  1. Fill COMPANY, TYPE, CATEGORY from Data_2026 payee lookup.
     Only fills rows where the field is currently empty AND the payee
     maps to exactly one value in Data_2026 history.

  2. Standardise BANK names:
       "TD"      → "TD - JMW"
       "TD Visa" → "TD - Visa"

Edits the April file IN PLACE. Reports every change made.

Usage:  python3 fill_lookup_fields.py <april_final.xlsx> <Data_2026.xlsx>
"""

import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

HEADER_ROW = 10
PAYEE_COL  = 9
COMPANY_COL= 8
TYPE_COL   = 10
CATEGORY_COL=11
BANK_COL   = 22

BANK_RENAMES = {
    "TD":      "TD - JMW",
    "TD Visa": "TD - Visa",
}


def build_lookup(data_path: Path) -> dict:
    """payee_upper → {field: single_value} for payees with exactly one value per field."""
    wb = load_workbook(data_path, read_only=True, data_only=True)
    ws = wb.active

    raw = defaultdict(lambda: {"COMPANY": set(), "TYPE": set(), "CATEGORY": set()})
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        payee = row[PAYEE_COL - 1]
        if payee is None:
            continue
        key = str(payee).strip().upper()
        for field, col in [("COMPANY", COMPANY_COL), ("TYPE", TYPE_COL), ("CATEGORY", CATEGORY_COL)]:
            val = row[col - 1]
            if val is not None:
                raw[key][field].add(str(val).strip())
    wb.close()

    # Keep only fields with exactly one distinct value
    lookup = {}
    for key, fields in raw.items():
        entry = {f: list(v)[0] for f, v in fields.items() if len(v) == 1}
        if entry:
            lookup[key] = entry
    return lookup


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 fill_lookup_fields.py <april_final.xlsx> <Data_2026.xlsx>")
        sys.exit(1)

    april_path = Path(sys.argv[1])
    data_path  = Path(sys.argv[2])

    print(f"Building lookup from: {data_path.name}")
    lookup = build_lookup(data_path)
    print(f"  {len(lookup)} payees with at least one exact-match field\n")

    wb = load_workbook(april_path)
    ws = wb.active

    filled   = defaultdict(int)   # field → rows filled
    renamed  = defaultdict(int)   # old_name → count
    no_match = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        payee_cell = ws.cell(row=row, column=PAYEE_COL)
        bank_cell  = ws.cell(row=row, column=BANK_COL)

        # Skip empty rows
        if payee_cell.value is None and bank_cell.value is None:
            continue

        # ── 1. Fill COMPANY / TYPE / CATEGORY ────────────────────────────────
        payee = payee_cell.value
        if payee is not None:
            key   = str(payee).strip().upper()
            entry = lookup.get(key, {})
            if not entry:
                no_match += 1
            for field, col in [("COMPANY", COMPANY_COL), ("TYPE", TYPE_COL), ("CATEGORY", CATEGORY_COL)]:
                cell = ws.cell(row=row, column=col)
                if cell.value is None and field in entry:
                    cell.value = entry[field]
                    filled[field] += 1

        # ── 2. Rename BANK ───────────────────────────────────────────────────
        if bank_cell.value is not None:
            old = str(bank_cell.value).strip()
            new = BANK_RENAMES.get(old)
            if new:
                bank_cell.value = new
                renamed[old] += 1

    wb.save(april_path)

    print("COMPANY / TYPE / CATEGORY fills:")
    for field in ["COMPANY", "TYPE", "CATEGORY"]:
        print(f"  {field:<12} filled in {filled[field]} rows")
    print(f"  Rows with no payee match in Data_2026: {no_match}")

    print("\nBANK renames:")
    for old, count in renamed.items():
        print(f"  {old!r:12} → {BANK_RENAMES[old]!r}  ({count} rows)")

    print(f"\nSaved: {april_path}")


if __name__ == "__main__":
    main()
