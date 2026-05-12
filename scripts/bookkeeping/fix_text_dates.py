"""
Fix rows where INVOICE DATE is stored as a text string (MM/DD/YYYY) instead of
a proper Excel date value. For each such row:

  1. Parse the text string → Python datetime
  2. Write it back as a proper Excel datetime so Excel recognises it as a date
  3. Recalculate WEEKENDS (Holiday / Weekend / Business Day) using the same
     BC + federal holiday calendar as fill_weekends.py
  4. Correct INVOICE QTR:  Jan-Mar → Q1, Apr-Jun → Q2, Jul-Sep → Q3, Oct-Dec → Q4
     (INVOICE YEAR stays 2026 for all rows in this file)

Fixes the file IN PLACE — no new filename suffix.

Usage:  python3 fix_text_dates.py <file.xlsx>
"""

import sys
from datetime import date, timedelta, datetime
from pathlib import Path
from openpyxl import load_workbook

SHEET       = "Sheet1"
HEADER_ROW  = 10
DATE_COL    = 5   # INVOICE DATE
WEEKENDS_COL= 7   # WEEKENDS
YEAR_COL    = 3   # INVOICE YEAR
QTR_COL     = 4   # INVOICE QTR


# --------------------------------------------------------------------------- #
# Holiday calendar (identical to fill_weekends.py)
# --------------------------------------------------------------------------- #
def easter_sunday(year: int) -> date:
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day   = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    first  = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (n - 1))


def bc_and_federal_holidays(year: int) -> dict:
    easter  = easter_sunday(year)
    may24   = date(year, 5, 24)
    vic_day = may24 - timedelta(days=may24.weekday() % 7)
    return {
        date(year, 1, 1):             "New Year's Day",
        nth_weekday(year, 2, 0, 3):   "Family Day",
        easter - timedelta(days=2):   "Good Friday",
        vic_day:                      "Victoria Day",
        date(year, 7, 1):             "Canada Day",
        nth_weekday(year, 8, 0, 1):   "BC Day",
        nth_weekday(year, 9, 0, 1):   "Labour Day",
        date(year, 9, 30):            "Truth and Reconciliation Day",
        nth_weekday(year, 10, 0, 2):  "Thanksgiving",
        date(year, 11, 11):           "Remembrance Day",
        date(year, 12, 25):           "Christmas Day",
        date(year, 12, 26):           "Boxing Day",
    }


def classify(d: date, holidays: dict) -> str:
    if d in holidays:
        return "Holiday"
    if d.weekday() >= 5:
        return "Weekend"
    return "Business Day"


def month_to_qtr(month: int) -> str:
    return f"Q{(month - 1) // 3 + 1}"


def parse_text_date(raw: str):
    """Parse MM/DD/YYYY text date. Returns a date object or None."""
    raw = raw.strip() if raw else ""
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    if len(sys.argv) != 2:
        print("Usage: python3 fix_text_dates.py <file.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    print(f"Opening: {path}")

    wb = load_workbook(path)
    ws = wb[SHEET]

    # Pre-build holiday calendars for years in the file
    holidays: dict[date, str] = {}
    for year in [2025, 2026, 2027]:
        holidays.update(bc_and_federal_holidays(year))

    fixed = 0
    skipped = 0

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=DATE_COL)
        val  = cell.value

        # Only process text-format dates
        if val is None or hasattr(val, "year"):
            continue

        raw = str(val).strip()
        d   = parse_text_date(raw)

        if d is None:
            print(f"  Row {row}: could not parse date {raw!r} — skipped")
            skipped += 1
            continue

        # 1. Fix the date cell → proper datetime (openpyxl stores dates as datetime)
        cell.value = datetime(d.year, d.month, d.day)

        # 2. Fix WEEKENDS
        ws.cell(row=row, column=WEEKENDS_COL).value = classify(d, holidays)

        # 3. Fix INVOICE QTR (YEAR stays 2026)
        ws.cell(row=row, column=QTR_COL).value  = month_to_qtr(d.month)
        ws.cell(row=row, column=YEAR_COL).value = d.year   # March rows → still 2026

        fixed += 1
        print(f"  Row {row}: {raw!r} → {d}  WEEKENDS={classify(d, holidays)}"
              f"  QTR={month_to_qtr(d.month)}")

    wb.save(path)
    print(f"\nFixed: {fixed} rows | Could not parse: {skipped} rows")
    print(f"Saved: {path}")


if __name__ == "__main__":
    main()
