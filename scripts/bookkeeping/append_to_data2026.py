"""
Append processed April rows into Data_2026.xlsx.
Writes a SAFE COPY: Data_2026_updated.xlsx — original is never modified.

Usage:  python3 append_to_data2026.py <april_final.xlsx> <Data_2026.xlsx>
Output: <Data_2026 directory>/Data_2026_updated.xlsx

Rules:
  - Copies only rows that have at least one non-None value.
  - Appends after the last row that has data in the CDN column (col U).
  - Preserves all cell values; does not copy formatting or formulas.
  - Reports row count before and after.
"""

import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

HEADER_ROW = 10
CDN_COL    = 21   # col U — used to find last populated row in Data_2026
TOTAL_COLS = 30   # columns B..AE (cols 2..30 in the schema) + col A = 30


def find_last_data_row(ws) -> int:
    last = HEADER_ROW
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False):
        for cell in row:
            if cell.value is not None:
                last = cell.row
                break
    return last


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 append_to_data2026.py <april_final.xlsx> <Data_2026.xlsx>")
        sys.exit(1)

    april_path = Path(sys.argv[1])
    data_path  = Path(sys.argv[2])
    out_path   = data_path.with_name("Data_2026_updated.xlsx")

    # Make a safe copy of Data_2026 — we write only to the copy
    print(f"Copying {data_path.name} → {out_path.name} ...")
    shutil.copy2(data_path, out_path)

    # Load April rows (read-only)
    print(f"Loading April data from: {april_path.name}")
    wb_april = load_workbook(april_path, read_only=True, data_only=True)
    ws_april = wb_april.active

    april_rows = []
    for row in ws_april.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if any(v is not None for v in row):
            april_rows.append(row)
    wb_april.close()
    print(f"  April rows to append: {len(april_rows)}")

    # Open the safe copy for writing
    print(f"Opening safe copy for writing: {out_path.name}")
    wb_dest = load_workbook(out_path)
    ws_dest = wb_dest.active

    first_empty = find_last_data_row(ws_dest) + 1
    print(f"  Last populated row in Data_2026: {first_empty - 1}")
    print(f"  Appending from row: {first_empty}")

    for i, row_vals in enumerate(april_rows):
        dest_row = first_empty + i
        for col_idx, val in enumerate(row_vals, start=1):
            if val is not None:
                ws_dest.cell(row=dest_row, column=col_idx, value=val)

    wb_dest.save(out_path)
    print(f"\nDone.")
    print(f"  Data_2026 rows before: {first_empty - 1 - HEADER_ROW}")
    print(f"  April rows appended:   {len(april_rows)}")
    print(f"  Total rows after:      {first_empty - 1 - HEADER_ROW + len(april_rows)}")
    print(f"  Safe copy written to:  {out_path}")
    print(f"\nReview {out_path.name} before renaming it to Data_2026.xlsx.")


if __name__ == "__main__":
    main()
