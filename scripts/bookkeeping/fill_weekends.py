"""
Fill column H ("WEEKENDS") in the Woodfine forensic reconstruction workbook.

Rule:
    Canadian/BC statutory holiday -> "Holiday"
    Saturday or Sunday            -> "Weekend"
    Otherwise                     -> "Business Day"

Date source:   column E ("INVOICE DATE")
Header row:    9
Data rows:     10 .. last populated row

Holidays covered (union of federal Canada Labour Code stats + BC Employment
Standards Act stats):
    New Year's Day        (Jan 1)
    Family Day            (3rd Monday of February, BC)
    Good Friday           (Friday before Easter Sunday)
    Victoria Day          (Monday on or before May 24)
    Canada Day            (Jul 1)
    BC Day                (1st Monday of August, BC)
    Labour Day            (1st Monday of September)
    Truth & Reconciliation (Sep 30)
    Thanksgiving          (2nd Monday of October)
    Remembrance Day       (Nov 11)
    Christmas Day         (Dec 25)
    Boxing Day            (Dec 26, federal stat)
"""

from datetime import date, timedelta
from pathlib import Path
from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Holiday calculation (pure stdlib; no `holidays` pkg needed)
# --------------------------------------------------------------------------- #
def easter_sunday(year: int) -> date:
    """Anonymous Gregorian algorithm (Meeus/Jones/Butcher)."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    """n-th `weekday` (Mon=0..Sun=6) of `month` in `year`."""
    first = date(year, month, 1)
    offset = (weekday - first.weekday()) % 7
    return first + timedelta(days=offset + 7 * (n - 1))


def bc_and_federal_holidays(year: int) -> dict[date, str]:
    easter = easter_sunday(year)
    may24 = date(year, 5, 24)
    victoria_day = may24 - timedelta(days=(may24.weekday()) % 7)  # Mon on/before May 24

    return {
        date(year, 1, 1):                      "New Year's Day",
        nth_weekday(year, 2, 0, 3):            "Family Day",
        easter - timedelta(days=2):            "Good Friday",
        victoria_day:                          "Victoria Day",
        date(year, 7, 1):                      "Canada Day",
        nth_weekday(year, 8, 0, 1):            "BC Day",
        nth_weekday(year, 9, 0, 1):            "Labour Day",
        date(year, 9, 30):                     "Truth and Reconciliation Day",
        nth_weekday(year, 10, 0, 2):           "Thanksgiving",
        date(year, 11, 11):                    "Remembrance Day",
        date(year, 12, 25):                    "Christmas Day",
        date(year, 12, 26):                    "Boxing Day",
    }


def classify(d: date, holidays: dict[date, str]) -> str:
    if d in holidays:
        return "Holiday"
    if d.weekday() >= 5:     # 5 = Sat, 6 = Sun
        return "Weekend"
    return "Business Day"


# --------------------------------------------------------------------------- #
# Fill the workbook
# --------------------------------------------------------------------------- #
SHEET       = "Sheet1"
HEADER_ROW  = 10
DATE_HEADER = "INVOICE DATE"   # resolved to a column by header lookup
FILL_HEADER = "WEEKENDS"       # resolved to a column by header lookup

OUTPUT_SUFFIX = "_WEEKENDS_FILLED"  # appended to the input stem for the output file


def find_column(ws, header_row: int, header_text: str) -> int:
    """Return the 1-based column index whose header (row `header_row`)
    matches `header_text` (case-insensitive, trimmed)."""
    target = header_text.strip().casefold()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if isinstance(v, str) and v.strip().casefold() == target:
            return col
    raise ValueError(
        f"Header {header_text!r} not found on row {header_row} of sheet {ws.title!r}"
    )


def prompt_for_input_path() -> Path:
    r"""Ask the user for the input .xlsx. Handles the usual Mac Terminal quirks:
    - paths wrapped in single or double quotes (copy/paste from Finder)
    - backslash-escaped spaces from drag-and-drop (e.g. /Users/jen/My\ File.xlsx)
    - leading '~' for home directory
    """
    while True:
        raw = input("Path to input .xlsx  (you can drag the file into this window): ").strip()
        if not raw:
            print("  -> empty path, try again.\n")
            continue

        # Strip one layer of wrapping quotes if present
        if (raw.startswith('"') and raw.endswith('"')) or \
           (raw.startswith("'") and raw.endswith("'")):
            raw = raw[1:-1]

        # Unescape backslash-escaped spaces (Mac drag-and-drop convention)
        raw = raw.replace("\\ ", " ")

        path = Path(raw).expanduser()

        if not path.exists():
            print(f"  -> file not found: {path}\n")
            continue
        if path.is_dir():
            print(f"  -> that's a folder, not a file: {path}\n")
            continue
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            print(f"  -> not an Excel file (.xlsx/.xlsm): {path}\n")
            continue

        return path


def main():
    in_path = prompt_for_input_path()
    out_path = in_path.with_name(in_path.stem + OUTPUT_SUFFIX + in_path.suffix)

    print(f"\nReading:  {in_path}")
    print(f"Writing:  {out_path}\n")

    wb = load_workbook(in_path)
    ws = wb[SHEET]

    date_col = find_column(ws, HEADER_ROW, DATE_HEADER)
    fill_col = find_column(ws, HEADER_ROW, FILL_HEADER)
    print(f"Resolved columns: {DATE_HEADER} -> col {date_col}, "
          f"{FILL_HEADER} -> col {fill_col}")

    # Pre-compute holidays for every year that appears in the data
    years_in_data = set()
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=date_col).value
        if hasattr(v, "year"):
            years_in_data.add(v.year)
    holidays: dict[date, str] = {}
    for y in years_in_data:
        holidays.update(bc_and_federal_holidays(y))

    # Walk and fill
    counts = {"Business Day": 0, "Weekend": 0, "Holiday": 0, "Skipped": 0}
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=date_col).value
        if not hasattr(v, "year"):
            counts["Skipped"] += 1
            continue
        d = v.date() if hasattr(v, "date") else v
        label = classify(d, holidays)
        ws.cell(row=row, column=fill_col).value = label
        counts[label] += 1

    wb.save(out_path)
    print(f"\nYears processed: {sorted(years_in_data)}")
    print(f"Holidays recognised this run:")
    for d in sorted(holidays):
        print(f"  {d}  ({d.strftime('%a')})  {holidays[d]}")
    print(f"Row counts: {counts}")
    print(f"\nWrote: {out_path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
    except Exception as e:
        print(f"\nERROR: {e}")
        raise
