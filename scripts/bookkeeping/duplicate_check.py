"""
Duplicate check before appending April data to Data_2026.
Composite key: (INVOICE_DATE, CDN_AMOUNT) — used because REF NUMBER is unpopulated.

Usage:  python3 duplicate_check.py <april_final.xlsx> <Data_2026.xlsx>
Exits 0 if clear; exits 1 if conflicts found (prints each conflict).
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

HEADER_ROW = 10
DATE_COL   = 5   # col E — INVOICE DATE
CDN_COL    = 21  # col U — CDN
PAYEE_COL  = 9   # col I — PAYEE (for conflict reporting)


def _normalize_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    return v


def load_key_set(path: Path) -> set:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    keys = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        d = _normalize_date(row[DATE_COL - 1])
        c = row[CDN_COL - 1]
        if d is not None and c is not None:
            keys.add((d, c))
    wb.close()
    return keys


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 duplicate_check.py <april_final.xlsx> <Data_2026.xlsx>")
        sys.exit(1)

    april_path = Path(sys.argv[1])
    data_path  = Path(sys.argv[2])

    print(f"Loading existing keys from: {data_path.name}")
    existing = load_key_set(data_path)
    print(f"  {len(existing)} unique (date, CDN) pairs in {data_path.name}")

    print(f"\nChecking April rows from: {april_path.name}")
    wb = load_workbook(april_path, read_only=True, data_only=True)
    ws = wb.active

    conflicts = []
    checked   = 0
    for row_num, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1
    ):
        d = _normalize_date(row[DATE_COL - 1])
        c = row[CDN_COL - 1]
        if d is None and c is None:
            continue
        checked += 1
        if (d, c) in existing:
            payee = row[PAYEE_COL - 1]
            conflicts.append((row_num, d, c, payee))
    wb.close()

    print(f"April rows checked: {checked}")

    if conflicts:
        print(f"\nCONFLICT: {len(conflicts)} row(s) in April already exist in {data_path.name}:")
        for row_num, d, c, payee in conflicts:
            print(f"  row {row_num}:  date={d}  CDN={c}  payee={payee}")
        print("\nAbort — resolve conflicts before appending.")
        sys.exit(1)

    print("Clear to merge — no duplicate (date, CDN) pairs found.")
    sys.exit(0)


if __name__ == "__main__":
    main()
