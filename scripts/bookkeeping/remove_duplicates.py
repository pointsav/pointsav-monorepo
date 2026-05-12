"""
Remove rows from the April FINAL file that already exist in Data_2026,
matched by (INVOICE_DATE, CDN_AMOUNT).

Deletes matching rows from the April file IN PLACE, then reports what was removed.

Usage:  python3 remove_duplicates.py <april_final.xlsx> <Data_2026.xlsx>
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

HEADER_ROW = 10
DATE_COL   = 5   # INVOICE DATE
CDN_COL    = 21  # CDN
PAYEE_COL  = 9   # PAYEE (for reporting)


def _normalize_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    return v


def load_key_set(path: Path) -> set:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    keys = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        d = _normalize_date(row[DATE_COL - 1])
        c = row[CDN_COL - 1]
        if d is not None and c is not None:
            keys.add((d, c))
    wb.close()
    return keys


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 remove_duplicates.py <april_final.xlsx> <Data_2026.xlsx>")
        sys.exit(1)

    april_path = Path(sys.argv[1])
    data_path  = Path(sys.argv[2])

    print(f"Loading existing keys from: {data_path.name}")
    existing = load_key_set(data_path)
    print(f"  {len(existing)} unique (date, CDN) pairs in {data_path.name}")

    # Find rows to delete (collect worksheet row numbers, scan bottom-to-top later)
    print(f"\nScanning: {april_path.name}")
    wb = load_workbook(april_path)
    ws = wb.active

    rows_to_delete = []
    for row_num in range(HEADER_ROW + 1, ws.max_row + 1):
        d = _normalize_date(ws.cell(row=row_num, column=DATE_COL).value)
        c = ws.cell(row=row_num, column=CDN_COL).value
        if d is None and c is None:
            continue
        if (d, c) in existing:
            payee = ws.cell(row=row_num, column=PAYEE_COL).value
            rows_to_delete.append((row_num, d, c, payee))

    if not rows_to_delete:
        print("No duplicates found — nothing to remove.")
        wb.close()
        return

    print(f"Removing {len(rows_to_delete)} duplicate rows:")
    for row_num, d, c, payee in rows_to_delete:
        print(f"  ws row {row_num}:  date={d}  CDN={c}  payee={payee}")

    # Delete from bottom to top so row-number shifts don't affect earlier deletes
    for row_num, _, _, _ in sorted(rows_to_delete, key=lambda x: x[0], reverse=True):
        ws.delete_rows(row_num)

    wb.save(april_path)
    print(f"\nRemoved {len(rows_to_delete)} rows. Saved: {april_path}")


if __name__ == "__main__":
    main()
