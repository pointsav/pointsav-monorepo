#!/usr/bin/env python3
import argparse
import re
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def process_excel(target_input, target_output, pattern):
    print(f"[SYSTEM] Engine: openpyxl (Native Excel)")
    try:
        wb = load_workbook(target_input)
    except Exception as e:
        print(f"[SYSTEM ERROR] Failed to load DOM: {e}")
        sys.exit(1)
        
    ws = wb.active
    idx_head = None
    
    # Locate header row ("REF NUMBER")
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if str(cell.value).strip() == "REF NUMBER":
                idx_head = cell.row
                break
        if idx_head:
            break
            
    if not idx_head:
        raise RuntimeError("Header locator 'REF NUMBER' missing.")

    # Map headers dynamically
    idx_map = {}
    for cell in ws[idx_head]:
        if cell.value:
            idx_map[str(cell.value).strip().upper()] = cell.column

    col_src = idx_map.get("PAYEE", 9)
    col_amt = idx_map.get("FORIGN CURRENCY AMOUNT", 18)
    col_cur = idx_map.get("CURENCY", 19)
    col_rate = idx_map.get("EXCHANGE RATE", 20)

    # Enforce target headers directly in the DOM
    ws.cell(row=idx_head, column=col_amt, value="FORIGN CURRENCY AMOUNT")
    ws.cell(row=idx_head, column=col_cur, value="Curency")
    ws.cell(row=idx_head, column=col_rate, value="Exchange rate")

    mutations = 0
    for r in range(idx_head + 1, ws.max_row + 1):
        src_cell = ws.cell(row=r, column=col_src)
        val = src_cell.value
        
        if isinstance(val, str):
            match = pattern.search(val)
            if match:
                ws.cell(row=r, column=col_amt, value=float(match.group("amt")))
                ws.cell(row=r, column=col_cur, value=match.group("cur"))
                
                rate_raw = match.group("rate")
                if rate_raw.startswith('.') or rate_raw.startswith('-.') or rate_raw.startswith('+.'):
                    rate_raw = rate_raw.replace('.', '0.', 1)
                ws.cell(row=r, column=col_rate, value=float(rate_raw))

                clean_val = (val[:match.start()] + val[match.end():]).strip()
                src_cell.value = re.sub(r"\s{2,}", " ", clean_val)
                mutations += 1

    print(f"[SYSTEM] Saving preserved DOM to: {target_output}")
    wb.save(target_output)
    return mutations

def process_csv(target_input, target_output, pattern):
    print(f"[SYSTEM] Engine: pandas (CSV DataFrame)")
    
    # Pre-scan to find the header row containing "REF NUMBER"
    df_raw = pd.read_csv(target_input, header=None, low_memory=False)
    header_idx = None
    for idx, row in df_raw.iterrows():
        if any(str(v).strip() == "REF NUMBER" for v in row.values if pd.notna(v)):
            header_idx = idx
            break
            
    if header_idx is None:
        raise RuntimeError("Header locator 'REF NUMBER' missing.")
        
    # Re-ingest properly locked onto the header
    df = pd.read_csv(target_input, skiprows=header_idx, low_memory=False)
    
    # Identify Payee Column dynamically
    payee_col = next((col for col in df.columns if str(col).strip().upper() == "PAYEE"), None)
    if not payee_col:
        raise RuntimeError("'PAYEE' column not found in schema.")

    # Ensure injection columns exist
    for col in ["FORIGN CURRENCY AMOUNT", "Curency", "Exchange rate"]:
        if col not in df.columns:
            df[col] = pd.NA

    mutations = 0
    for idx, row in df.iterrows():
        val = str(row[payee_col])
        if val and val.lower() != 'nan':
            match = pattern.search(val)
            if match:
                df.at[idx, "FORIGN CURRENCY AMOUNT"] = float(match.group("amt"))
                df.at[idx, "Curency"] = match.group("cur")
                
                rate_raw = match.group("rate")
                if rate_raw.startswith('.') or rate_raw.startswith('-.') or rate_raw.startswith('+.'):
                    rate_raw = rate_raw.replace('.', '0.', 1)
                df.at[idx, "Exchange rate"] = float(rate_raw)
                
                clean_val = (val[:match.start()] + val[match.end():]).strip()
                df.at[idx, payee_col] = re.sub(r"\s{2,}", " ", clean_val)
                mutations += 1

    print(f"[SYSTEM] Saving processed dataset to: {target_output}")
    df.to_csv(target_output, index=False)
    return mutations

def execute_extraction():
    parser = argparse.ArgumentParser(description="Extract foreign currency strings from PAYEE column.")
    parser.add_argument("--input",  help="Direct path to input .xlsx or .csv (bypasses inbox scan)")
    parser.add_argument("--output", help="Output file path (default: <input>_FINAL<ext>)")
    args = parser.parse_args()

    if args.input:
        target_input = Path(args.input)
        if not target_input.exists():
            print(f"[SYSTEM ERROR] File not found: {target_input}")
            sys.exit(1)
        target_output = Path(args.output) if args.output else \
            target_input.with_name(f"{target_input.stem}_FINAL{target_input.suffix}")
        print(f"[SYSTEM] Direct mode. Target payload: {target_input.name}")
    else:
        inbox_dir = Path("/Users/Office/Foundry/inbox")

        if not inbox_dir.exists():
            print(f"[SYSTEM ERROR] Payload directory missing: {inbox_dir}")
            sys.exit(1)

        # 1. Directory Sweeping & Exclusion Protocol for both formats
        valid_files = []
        for ext in ["*.xlsx", "*.csv"]:
            for file_path in inbox_dir.glob(ext):
                name = file_path.name
                if "_FINAL" not in name and "~$" not in name:
                    valid_files.append(file_path)

        # 2. Singularity Lock
        if len(valid_files) == 0:
            print("[SYSTEM ABORT] No valid payload files (.xlsx or .csv) detected in /inbox/")
            sys.exit(1)
        elif len(valid_files) > 1:
            print("[SYSTEM ABORT] Multiple payload files detected. Singularity lock failed to prevent cross-contamination.")
            for f in valid_files:
                print(f" - {f.name}")
            sys.exit(1)

        target_input = valid_files[0]
        target_output = target_input.with_name(f"{target_input.stem}_FINAL{target_input.suffix}")
        print(f"[SYSTEM] Singularity Lock Engaged. Target payload: {target_input.name}")

    # Standard Regex Pattern
    pattern = re.compile(r"(?P<amt>[-+]?\d*\.?\d+)\s*(?P<cur>[A-Z]{3})\s*@\s*(?P<rate>[-+]?\d*\.?\d+)")

    try:
        if target_input.suffix.lower() == '.xlsx':
            mutations = process_excel(target_input, target_output, pattern)
        elif target_input.suffix.lower() == '.csv':
            mutations = process_csv(target_input, target_output, pattern)
        else:
            print(f"[SYSTEM ERROR] Unsupported file format: {target_input.suffix}")
            sys.exit(1)
            
        print(f"[SYSTEM] Extraction complete. Payee rows mutated: {mutations}")
        
    except Exception as e:
        print(f"[SYSTEM FATAL ERROR] {e}")
        sys.exit(1)

if __name__ == "__main__":
    execute_extraction()
