"""
Generate a monthly summary CSV for a processed transactions file.

Reports:
  - Total rows
  - Sum of CDN amounts (total, debit, credit)
  - Row count and CDN sum by BANK
  - WEEKENDS distribution
  - FX transaction count (rows with a foreign currency amount)
  - Count by currency code
  - Rows missing key fields (INVOICE DATE, CDN, PAYEE)

Usage:  python3 monthly_summary.py <input.xlsx> [output.csv]
Default output: <stem>_SUMMARY.csv
"""

import sys
import csv
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

HEADER_ROW  = 10
DATE_COL    = 5   # INVOICE DATE
WEEKENDS_COL= 7   # WEEKENDS
PAYEE_COL   = 9   # PAYEE
CDN_COL     = 21  # CDN
BANK_COL    = 22  # BANK
FX_AMT_COL  = 18  # FORIGN CURRENCY AMOUNT
FX_CUR_COL  = 19  # Curency


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 monthly_summary.py <input.xlsx> [output.csv]")
        sys.exit(1)

    in_path  = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else \
               in_path.with_name(in_path.stem + "_SUMMARY.csv")

    wb = load_workbook(in_path, read_only=True, data_only=True)
    ws = wb.active

    total_rows   = 0
    total_cdn    = 0.0
    total_debit  = 0.0
    total_credit = 0.0
    bank_counts  = defaultdict(int)
    bank_cdn     = defaultdict(float)
    weekends_dist = defaultdict(int)
    fx_count     = 0
    currency_counts = defaultdict(int)
    missing_date = 0
    missing_cdn  = 0
    missing_payee = 0

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        # Skip fully empty rows
        if not any(v is not None for v in row):
            continue

        total_rows += 1

        date_val    = row[DATE_COL    - 1]
        weekends_val= row[WEEKENDS_COL- 1]
        payee_val   = row[PAYEE_COL   - 1]
        cdn_val     = row[CDN_COL     - 1]
        bank_val    = row[BANK_COL    - 1]
        fx_amt      = row[FX_AMT_COL  - 1]
        fx_cur      = row[FX_CUR_COL  - 1]

        if date_val is None:
            missing_date += 1
        if payee_val is None:
            missing_payee += 1
        if cdn_val is None:
            missing_cdn += 1

        if cdn_val is not None:
            try:
                amount = float(cdn_val)
                total_cdn += amount
                if amount >= 0:
                    total_debit += amount
                else:
                    total_credit += amount
            except (TypeError, ValueError):
                pass

        bank_key = str(bank_val).strip() if bank_val else "(none)"
        bank_counts[bank_key] += 1
        if cdn_val is not None:
            try:
                bank_cdn[bank_key] += float(cdn_val)
            except (TypeError, ValueError):
                pass

        if weekends_val:
            weekends_dist[str(weekends_val)] += 1
        else:
            weekends_dist["(not set)"] += 1

        if fx_amt is not None:
            fx_count += 1
            cur_key = str(fx_cur).strip() if fx_cur else "(unknown)"
            currency_counts[cur_key] += 1

    wb.close()

    rows = []
    rows.append(["SECTION", "LABEL", "VALUE"])
    rows.append(["", "", ""])

    rows.append(["TOTALS", "Total transaction rows", total_rows])
    rows.append(["TOTALS", "Total CDN amount",       f"{total_cdn:,.2f}"])
    rows.append(["TOTALS", "Total debits (>=0)",     f"{total_debit:,.2f}"])
    rows.append(["TOTALS", "Total credits (<0)",     f"{total_credit:,.2f}"])
    rows.append(["", "", ""])

    rows.append(["BY BANK", "Bank", "Row count | CDN sum"])
    for bank in sorted(bank_counts):
        rows.append(["BY BANK", bank, f"{bank_counts[bank]} rows | {bank_cdn[bank]:,.2f} CDN"])
    rows.append(["", "", ""])

    rows.append(["WEEKENDS", "Classification", "Count"])
    for label in sorted(weekends_dist):
        rows.append(["WEEKENDS", label, weekends_dist[label]])
    rows.append(["", "", ""])

    rows.append(["FX", "Foreign currency rows", fx_count])
    for cur in sorted(currency_counts):
        rows.append(["FX", f"  {cur}", currency_counts[cur]])
    rows.append(["", "", ""])

    rows.append(["DATA QUALITY", "Rows missing INVOICE DATE", missing_date])
    rows.append(["DATA QUALITY", "Rows missing CDN amount",   missing_cdn])
    rows.append(["DATA QUALITY", "Rows missing PAYEE",        missing_payee])

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    print(f"Summary written to: {out_path}")
    print(f"\nQuick view:")
    print(f"  Total rows:    {total_rows}")
    print(f"  Total CDN:     {total_cdn:,.2f}")
    print(f"  Debits:        {total_debit:,.2f}  |  Credits: {total_credit:,.2f}")
    print(f"  FX rows:       {fx_count}")
    print(f"  Missing date:  {missing_date}  |  Missing CDN: {missing_cdn}")


if __name__ == "__main__":
    main()
